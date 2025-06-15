from typing import List, Dict, Any
from langchain.schema import Document
from langchain.text_splitter import RecursiveCharacterTextSplitter
import openpyxl.utils
import pdfplumber
import pandas as pd
import logging
import openpyxl

logger = logging.getLogger(__name__)

class Pdf():
    def __init__(self, pdf_path: str = None):
        self.pdf_path = pdf_path
        self.documents = []
        self.tables = []
        self.text_content = ""

    def extract_with_pdfplumber(self, pdf_path: str = None) -> Dict[str, List[Document]]:
        """
        使用pdfplumber提取PDF的所有内容（文本和表格）
        
        Args:
            pdf_path: PDF文件路径
        
        Returns:
            Dict包含text_elements和table_elements列表
        """
        path = pdf_path or self.pdf_path
        if not path:
            raise ValueError("PDF 路径不能为空")
            
        text_elements = []
        table_elements = []

        try:
            with pdfplumber.open(path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    
                    # 1. 提取页面中的表格
                    page_tables = page.extract_tables()
                    table_regions = []
                    
                    # 处理表格并记录表格区域
                    if page_tables:
                        for table_idx, table in enumerate(page_tables):
                            if table and len(table) > 1:  # 确保表格有数据
                                try:
                                    # 创建DataFrame
                                    headers = table[0] if table[0] else [f"Col_{i}" for i in range(len(table[1]))]
                                    df = pd.DataFrame(table[1:], columns=headers)
                                    
                                    # 清理空行和空列
                                    df = df.dropna(how='all').dropna(axis=1, how='all')
                                    
                                    if not df.empty:
                                        table_doc = Document(
                                            page_content=df.to_markdown(index=False),
                                            metadata={
                                                "source": path,
                                                "type": "table",
                                                "page": page_num + 1,
                                                "table_idx": table_idx,
                                                "extracted_by": "pdfplumber"
                                            }
                                        )
                                        table_elements.append(table_doc)
                                        
                                        # 尝试获取表格在页面中的大致位置
                                        try:
                                            # 简单的表格区域估算方法
                                            first_cell = next((cell for row in table for cell in row if cell and str(cell).strip()), None)
                                            if first_cell:
                                                # 在页面中搜索第一个单元格的位置
                                                search_text = str(first_cell).strip()[:20]  # 只取前20个字符
                                                for char in page.chars:
                                                    if search_text in char.get('text', ''):
                                                        # 估算表格区域
                                                        table_region = {
                                                            'x0': max(0, char['x0'] - 50),
                                                            'top': max(0, char['top'] - 20),
                                                            'x1': min(page.width, char['x0'] + 300),
                                                            'bottom': min(page.height, char['top'] + len(table) * 15)
                                                        }
                                                        table_regions.append(table_region)
                                                        break
                                        except:
                                            pass  # 如果无法定位表格位置，继续处理
                                            
                                except Exception as e:
                                    logger.warning(f"处理第{page_num+1}页表格{table_idx}时出错: {e}")
                    
                    # 2. 提取文本内容，尽量避开表格区域
                    try:
                        if table_regions:
                            # 如果有表格，尝试过滤表格区域的文本
                            filtered_chars = []
                            for char in page.chars:
                                char_in_table = False
                                for region in table_regions:
                                    if (region['x0'] <= char['x0'] <= region['x1'] and 
                                        region['top'] <= char['top'] <= region['bottom']):
                                        char_in_table = True
                                        break
                                if not char_in_table:
                                    filtered_chars.append(char)
                            
                            # 重建文本
                            if filtered_chars:
                                filtered_chars.sort(key=lambda x: (x['top'], x['x0']))
                                text = ''.join(char['text'] for char in filtered_chars)
                            else:
                                text = ""
                        else:
                            # 没有表格，直接提取文本
                            text = page.extract_text()
                        
                        # 清理文本
                        if text and text.strip():
                            # 移除明显的表格残留和多余空白
                            lines = text.split('\n')
                            clean_lines = []
                            
                            for line in lines:
                                line = line.strip()
                                if line:
                                    # 跳过看起来像表格分隔符的行
                                    if not all(c in '|-+=_ \t' for c in line):
                                        # 跳过包含过多表格分隔符的行
                                        if line.count('|') <= 3:  # 允许一些管道符
                                            clean_lines.append(line)
                            
                            if clean_lines:
                                clean_content = '\n'.join(clean_lines)
                                text_doc = Document(
                                    page_content=clean_content,
                                    metadata={
                                        "source": path,
                                        "type": "text",
                                        "page": page_num + 1,
                                        "tables_filtered": len(page_tables) if page_tables else 0
                                    }
                                )
                                text_elements.append(text_doc)
                    
                    except Exception as e:
                        logger.warning(f"处理第{page_num+1}页文本时出错: {e}")
            
            logger.info(f"pdfplumber提取完成: {len(text_elements)} 个文本元素, {len(table_elements)} 个表格")
            
            # 更新实例状态
            self.tables = table_elements
            
            return {
                "text_elements": text_elements,
                "table_elements": table_elements
            }
            
        except Exception as e:
            logger.error(f"pdfplumber 提取失败：{e}")
            return {
                "text_elements": [],
                "table_elements": []
            }

    def comprehensive_extract(self, pdf_path: str = None) -> Dict[str, Any]:
        """
        综合提取PDF内容（简化版，只使用pdfplumber）
        """
        result = self.extract_with_pdfplumber(pdf_path)
        
        return {
            'text_elements': result['text_elements'],
            'table_elements': result['table_elements'],
            'total_tables': len(result['table_elements'])
        }
    
    def tables_to_text(self, tables: List[pd.DataFrame]) -> str:
        text_parts = []

        for table in tables:
            table_name = getattr(table, 'name', f"Table_{len(text_parts)+1}")
            text_parts.append(f"\n=== {table_name} ===")

            markdown_table = table.to_markdown(index=False)
            text_parts.append(markdown_table)
            text_parts.append("") # 换行

        return "\n".join(text_parts)

def chunk(file_path: str, chunk_size: int = 200, chunk_overlap: int = 50) -> List[Document]:
    """
    处理PDF文件，提取文本和表格内容，并分块返回Document列表
    
    Args:
        file_path: PDF文件路径
        chunk_size: 分块大小，默认300字符
        chunk_overlap: 分块重叠大小，默认100字符
    
    Returns:
        List[Document]: 分块后的文档列表
    """
    try:
        # 创建PDF处理器
        pdf = Pdf(file_path)

        # 使用简化的提取方法
        extraction_result = pdf.comprehensive_extract()
        
        # 使用文本分割器
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=chunk_size,
            chunk_overlap=chunk_overlap,
            separators=["\n\n", "\n", "。", "！", "？", " ", ""]
        )
        
        documents = []
        
        # 1. 处理文本元素
        for text_element in extraction_result['text_elements']:
            content = text_element.page_content.strip()
            if content:
                if len(content) > chunk_size:
                    # 分块
                    chunks = text_splitter.split_text(content)
                    for chunk_idx, chunk_text in enumerate(chunks):
                        chunk_doc = Document(
                            page_content=chunk_text,
                            metadata={
                                **text_element.metadata,
                                "chunk_index": chunk_idx,
                                "total_chunks": len(chunks),
                                "is_chunked": True
                            }
                        )
                        documents.append(chunk_doc)
                else:
                    # 直接使用
                    text_element.metadata["is_chunked"] = False
                    documents.append(text_element)
        
        # 2. 表格元素直接添加，不分块
        documents.extend(extraction_result["table_elements"])
        
        logger.info(f"PDF处理完成: {file_path}, 文档数量: {len(documents)}")
        
        return documents
        
    except Exception as e:
        logger.error(f"PDF分块处理失败: {file_path}, 错误: {e}")
        error_doc = Document(
            page_content=f"PDF处理失败: {str(e)}",
            metadata={
                "source": file_path,
                "type": "error",
                "error": str(e)
            }
        )
        return [error_doc]


def read_and_process_excel(file_path):
    data = []
    # 打开 Excel 文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # 创建一个字典来存储单元格值
    cell_values = {}

    # 遍历所有单元格
    for row in sheet.iter_rows():
        for cell in row:
            cell_values[cell.coordinate] = cell.value

    # 遍历所有合并单元格
    for merged_range in sheet.merged_cells.ranges:
        min_col, min_row = merged_range.min_col, merged_range.min_row
        max_col, max_row = merged_range.max_col, merged_range.max_row
        
        # 获取合并单元格的值（通常位于合并区域的左上角）
        value = sheet.cell(row=min_row, column=min_col).value
        
        # 将值填充到所有合并区域的单元格中
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_coordinate = openpyxl.utils.get_column_letter(col) + str(row)
                cell_values[cell_coordinate] = value

    # 输出所有单元格的值
    max_row = sheet.max_row
    max_col = sheet.max_column
    for row in range(1, max_row + 1):
        row_values = []
        for col in range(1, max_col + 1):
            cell_coordinate = openpyxl.utils.get_column_letter(col) + str(row)
            cell_value = cell_values.get(cell_coordinate, None)
            row_values.append(str(cell_value).replace('\n', '') if cell_value is not None else '')
        data.append(row_values)
    return data