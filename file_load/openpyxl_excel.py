from openpyxl import load_workbook

wb = load_workbook(filename="./file_load/fixtures/zhidu_detail.xlsx")
ws = wb.active

# 打印每一行
for row in ws.iter_rows():
    info = []
    for cell in row:
        info.append(cell.value)
    print(len(info), info)

# 获取合并单元格的值
for merged_range in ws.merged_cells.ranges:
    value = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    print(merged_range, value)
    break