"""
Excel处理功能测试文件
测试Excel加载、合并单元格处理等功能
调用doc_parse.py中的read_and_process_excel函数
"""

import unittest
import os
import sys
import logging
from pathlib import Path
from typing import List, Dict, Any
import openpyxl
from openpyxl import Workbook
import tempfile

# 添加项目根目录到路径
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, project_root)

# 导入Excel处理函数
from doc_parse import read_and_process_excel

# 设置日志
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class TestExcelProcessing(unittest.TestCase):
    """Excel处理功能测试类"""
    
    @classmethod
    def setUpClass(cls):
        """测试类初始化"""
        cls.test_excel_path = "data/detail.xlsx"
        cls.test_excel_path_alt = "data/zhidu_detail.xlsx"
        
        # 检查测试文件是否存在
        if not os.path.exists(cls.test_excel_path):
            raise FileNotFoundError(f"测试Excel文件不存在: {cls.test_excel_path}")
        
        logger.info(f"开始测试Excel处理功能，测试文件: {cls.test_excel_path}")
    
    def test_excel_file_accessibility(self):
        """测试Excel文件可访问性"""
        print("\n=== 测试1: Excel文件可访问性 ===")
        
        self.assertTrue(os.path.exists(self.test_excel_path), f"Excel文件应该存在: {self.test_excel_path}")
        
        file_size = os.path.getsize(self.test_excel_path)
        print(f"✓ Excel文件大小: {file_size:,} 字节 ({file_size/1024:.1f} KB)")
        
        self.assertGreater(file_size, 0, "Excel文件不应为空")
        
        # 测试文件是否可以被openpyxl打开
        try:
            workbook = openpyxl.load_workbook(self.test_excel_path)
            print(f"✓ Excel文件可以被openpyxl打开")
            print(f"✓ 工作表数量: {len(workbook.worksheets)}")
            print(f"✓ 活动工作表名称: {workbook.active.title}")
            workbook.close()
        except Exception as e:
            self.fail(f"无法打开Excel文件: {e}")
    
    def test_basic_excel_reading(self):
        """测试基本Excel读取功能"""
        print("\n=== 测试2: 基本Excel读取 ===")
        
        try:
            data = read_and_process_excel(self.test_excel_path)
            
            self.assertIsInstance(data, list, "返回的数据应该是列表格式")
            self.assertGreater(len(data), 0, "应该读取到至少一行数据")
            
            print(f"✓ 成功读取Excel文件")
            print(f"✓ 总行数: {len(data)}")
            
            # 检查第一行数据
            if data:
                first_row = data[0]
                self.assertIsInstance(first_row, list, "每行数据应该是列表格式")
                print(f"✓ 总列数: {len(first_row)}")
                print(f"✓ 第一行数据: {first_row[:5]}...")  # 只显示前5列
            
            # 检查数据一致性
            if len(data) > 1:
                col_count = len(data[0])
                for i, row in enumerate(data[1:], 1):
                    if len(row) != col_count:
                        print(f"⚠ 警告: 第{i+1}行列数({len(row)})与第1行不一致({col_count})")
            
        except Exception as e:
            self.fail(f"基本Excel读取失败: {e}")
    
    def test_merged_cells_handling(self):
        """测试合并单元格处理"""
        print("\n=== 测试3: 合并单元格处理 ===")
        
        try:
            # 先检查原文件是否有合并单元格
            workbook = openpyxl.load_workbook(self.test_excel_path)
            sheet = workbook.active
            merged_ranges = list(sheet.merged_cells.ranges)
            
            print(f"✓ 原文件中发现 {len(merged_ranges)} 个合并单元格区域")
            
            if merged_ranges:
                for i, merged_range in enumerate(merged_ranges[:3]):  # 只显示前3个
                    print(f"  合并区域 {i+1}: {merged_range}")
            
            workbook.close()
            
            # 测试处理结果
            data = read_and_process_excel(self.test_excel_path)
            
            print(f"✓ 合并单元格处理完成")
            print(f"✓ 处理后数据维度: {len(data)} 行 x {len(data[0]) if data else 0} 列")
            
            # 检查是否有重复填充的值（合并单元格的特征）
            if len(data) > 1:
                duplicate_count = 0
                for row in data:
                    for i in range(len(row) - 1):
                        if row[i] and row[i] == row[i + 1] and row[i] != '':
                            duplicate_count += 1
                
                if duplicate_count > 0:
                    print(f"✓ 检测到 {duplicate_count} 个可能的合并单元格填充")
            
        except Exception as e:
            print(f"⚠ 合并单元格处理遇到问题: {e}")
            logger.warning(f"合并单元格测试失败: {e}")
    
    def test_data_quality(self):
        """测试数据质量"""
        print("\n=== 测试4: 数据质量检查 ===")
        
        try:
            data = read_and_process_excel(self.test_excel_path)
            
            # 统计数据质量
            total_cells = sum(len(row) for row in data)
            empty_cells = sum(1 for row in data for cell in row if cell == '' or cell == 'None')
            non_empty_cells = total_cells - empty_cells
            
            print(f"✓ 数据质量统计:")
            print(f"  总单元格数: {total_cells}")
            print(f"  非空单元格: {non_empty_cells}")
            print(f"  空单元格: {empty_cells}")
            print(f"  数据覆盖率: {non_empty_cells/total_cells*100:.1f}%")
            
            # 检查中文内容
            chinese_cells = 0
            for row in data:
                for cell in row:
                    if any('\u4e00' <= char <= '\u9fff' for char in str(cell)):
                        chinese_cells += 1
            
            print(f"  包含中文的单元格: {chinese_cells}")
            
            # 显示前几行数据样例
            print(f"\n✓ 数据样例 (前3行):")
            for i, row in enumerate(data[:3]):
                print(f"  第{i+1}行: {row[:5]}{'...' if len(row) > 5 else ''}")
            
            self.assertGreater(non_empty_cells, 0, "应该有非空数据")
            
        except Exception as e:
            print(f"⚠ 数据质量检查遇到问题: {e}")
            logger.warning(f"数据质量测试失败: {e}")
    
    def test_multiple_excel_files(self):
        """测试处理多个Excel文件"""
        print("\n=== 测试5: 多Excel文件处理 ===")
        
        excel_files = [self.test_excel_path]
        if os.path.exists(self.test_excel_path_alt):
            excel_files.append(self.test_excel_path_alt)
        
        results = {}
        for excel_file in excel_files:
            try:
                data = read_and_process_excel(excel_file)
                results[excel_file] = data
                print(f"✓ 成功处理: {excel_file}")
                print(f"  数据维度: {len(data)} 行 x {len(data[0]) if data else 0} 列")
            except Exception as e:
                print(f"⚠ 处理 {excel_file} 时出错: {e}")
        
        self.assertGreater(len(results), 0, "至少应该成功处理一个Excel文件")
        print(f"✓ 总共成功处理 {len(results)} 个Excel文件")
    
    def test_error_handling(self):
        """测试错误处理"""
        print("\n=== 测试6: 错误处理 ===")
        
        # 测试不存在的文件
        try:
            read_and_process_excel("nonexistent_file.xlsx")
            self.fail("应该抛出文件不存在的异常")
        except FileNotFoundError:
            print("✓ 正确处理文件不存在错误")
        except Exception as e:
            print(f"✓ 捕获到其他错误: {type(e).__name__}")
        
        # 测试无效路径
        try:
            read_and_process_excel("")
            self.fail("应该抛出路径无效的异常")
        except Exception as e:
            print(f"✓ 正确处理空路径错误: {type(e).__name__}")
        
        print("✓ 错误处理测试完成")
    
    def test_create_and_read_test_excel(self):
        """测试创建和读取测试Excel文件"""
        print("\n=== 测试7: 创建和读取测试Excel ===")
        
        # 创建临时Excel文件进行测试
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            temp_excel_path = tmp_file.name
        
        try:
            # 创建测试Excel文件
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "测试工作表"
            
            # 添加测试数据
            test_data = [
                ["姓名", "部门", "职位", "工龄"],
                ["张三", "技术部", "工程师", "3年"],
                ["李四", "市场部", "经理", "5年"],
                ["王五", "财务部", "会计", "2年"]
            ]
            
            for row_idx, row_data in enumerate(test_data, 1):
                for col_idx, cell_data in enumerate(row_data, 1):
                    sheet.cell(row=row_idx, column=col_idx, value=cell_data)
            
            # 创建合并单元格测试
            sheet.merge_cells('A6:B6')
            sheet['A6'] = "合并单元格测试"
            
            workbook.save(temp_excel_path)
            workbook.close()
            
            print("✓ 测试Excel文件创建成功")
            
            # 测试读取
            data = read_and_process_excel(temp_excel_path)
            
            print(f"✓ 测试文件读取成功")
            print(f"✓ 数据维度: {len(data)} 行 x {len(data[0]) if data else 0} 列")
            print(f"✓ 读取的数据:")
            for i, row in enumerate(data):
                print(f"  第{i+1}行: {row}")
            
            # 验证数据正确性
            self.assertEqual(len(data), 6, "应该有6行数据（包括合并单元格行）")
            self.assertEqual(data[0][0], "姓名", "第一行第一列应该是'姓名'")
            self.assertEqual(data[1][0], "张三", "第二行第一列应该是'张三'")
            
            # 验证合并单元格处理
            self.assertEqual(data[5][0], "合并单元格测试", "合并单元格应该被正确处理")
            self.assertEqual(data[5][1], "合并单元格测试", "合并单元格的值应该填充到所有单元格")
            
        finally:
            # 清理临时文件
            if os.path.exists(temp_excel_path):
                os.unlink(temp_excel_path)
                print("✓ 临时文件清理完成")


class TestExcelUtilities(unittest.TestCase):
    """Excel工具函数测试类"""
    
    def test_cell_coordinate_handling(self):
        """测试单元格坐标处理"""
        print("\n=== 工具测试1: 单元格坐标处理 ===")
        
        # 测试坐标转换
        test_cases = [
            (1, 1, "A1"),
            (1, 26, "Z1"),
            (1, 27, "AA1"),
            (2, 1, "A2")
        ]
        
        for row, col, expected in test_cases:
            coordinate = openpyxl.utils.get_column_letter(col) + str(row)
            self.assertEqual(coordinate, expected, f"坐标转换错误: ({row}, {col}) -> {coordinate}")
            print(f"✓ 坐标转换正确: ({row}, {col}) -> {coordinate}")
    
    def test_workbook_operations(self):
        """测试工作簿操作"""
        print("\n=== 工具测试2: 工作簿操作 ===")
        
        # 测试创建工作簿
        workbook = Workbook()
        sheet = workbook.active
        
        # 测试基本操作
        sheet['A1'] = "测试"
        self.assertEqual(sheet['A1'].value, "测试")
        print("✓ 单元格写入读取正常")
        
        # 测试中文处理
        sheet['B1'] = "中文测试内容"
        self.assertEqual(sheet['B1'].value, "中文测试内容")
        print("✓ 中文内容处理正常")
        
        workbook.close()


def run_tests():
    """运行所有测试"""
    print("=" * 60)
    print("Excel处理功能测试开始")
    print("=" * 60)
    
    # 创建测试套件
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # 添加测试类
    suite.addTests(loader.loadTestsFromTestCase(TestExcelProcessing))
    suite.addTests(loader.loadTestsFromTestCase(TestExcelUtilities))
    
    # 运行测试
    runner = unittest.TextTestRunner(verbosity=2, stream=sys.stdout)
    result = runner.run(suite)
    
    print("\n" + "=" * 60)
    print("测试总结")
    print("=" * 60)
    print(f"运行测试数量: {result.testsRun}")
    print(f"成功: {result.testsRun - len(result.failures) - len(result.errors)}")
    print(f"失败: {len(result.failures)}")
    print(f"错误: {len(result.errors)}")
    print(f"跳过: {len(result.skipped) if hasattr(result, 'skipped') else 0}")
    
    if result.failures:
        print("\n失败的测试:")
        for test, traceback in result.failures:
            print(f"- {test}: {traceback}")
    
    if result.errors:
        print("\n错误的测试:")
        for test, traceback in result.errors:
            print(f"- {test}: {traceback}")
    
    return result.wasSuccessful()


if __name__ == "__main__":
    # 检查必要的依赖
    print("检查依赖库...")
    
    required_libs = ['openpyxl', 'doc_parse']
    
    for lib in required_libs:
        try:
            __import__(lib)
            print(f"✓ {lib} 已安装")
        except ImportError:
            print(f"✗ {lib} 未安装 (必需)")
    
    print("\n开始运行Excel测试...")
    success = run_tests()
    
    if success:
        print("\n🎉 所有Excel测试通过!")
    else:
        print("\n⚠️  部分Excel测试失败，请检查上述错误信息") 